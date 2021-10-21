import openpyxl

from utils.file_reader import FileReader


class SynonymsExcel:
    def __init__(self, filepath) -> None:
        self.wb = openpyxl.load_workbook(filepath)
        print(self.wb.get_sheet_names())
        self.sheet = self.wb.get_sheet_by_name('Таблица синонимов')
        print(self.wb.active)

    def get_synonyms(self, num_of_cols):
        synonyms = dict()
        for column in range(1, num_of_cols):
            key_synonym = self.sheet.cell(row=1,column=column).value
            synonyms[key_synonym] = [key_synonym]
            counter = 2
            while True:
                synonym = self.sheet.cell(row=counter,column=column).value
                if not synonym:
                    break
                synonyms[key_synonym].append(synonym)
                counter+=1
        return synonyms




from utils.text_parser import TextParser
if __name__ == '__main__':
    ex = SynonymsExcel('../synonyms/synonyms_lexems.xlsx')
    parser = TextParser()
    synonyms = ex.get_synonyms(38)
    in_txt = FileReader('../resources/fullrequests.txt')
    lexems = parser.get_lexems_from_text(in_txt.read_lines())
    new_lexems = parser.replace_to_synonym(lexems, synonyms)
    print(synonyms)